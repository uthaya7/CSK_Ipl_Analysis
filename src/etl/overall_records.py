import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def combine_all_records():
    # File paths
    base_path = r"F:\Data Analytics\Projects\csk_analysis\data\final_excel"
    batting_path = base_path + r"\batting_records_csk_cleaned.xlsx"
    bowling_path = base_path + r"\bowling_records_csk_cleaned.xlsx"
    fielding_path = base_path + r"\fielding_records_csk_cleaned.xlsx"

    output_excel = base_path + r"\All_Players_Records.xlsx"

    print(f"--- Combining Records into {output_excel.split(r'\\')[-1]} ---")

    # Helper to read all sheets
    def read_all_sheets(file_path, dtype=None):
        excel_file = pd.ExcelFile(file_path)
        dfs = []
        for sheet in excel_file.sheet_names:
            # Read the sheet
            df = pd.read_excel(file_path, sheet_name=sheet, dtype=dtype)
            df.columns = df.columns.str.strip().str.replace(' ', '_', regex=False) # Clean column names
            # No need to add Season here, it's already in the cleaned files
            dfs.append(df)
        return pd.concat(dfs, ignore_index=True)

    # ----------------------------- 🟡 1. BATTING -----------------------------
    print("Processing Batting data...")
    df_bat = read_all_sheets(batting_path)

    # Explicitly define columns and ensure they exist, using 0 as default for numeric
    bat_cols = ['Player_ID', 'Player', 'Full_Name', 'Season', 'Mat', 'Inns', 'NO', 'Runs', 'HS', 'HS_Numeric',
                'Not_out_status', 'Ave', 'BF', 'SR', 'Fours', 'Sixes', 'Ducks', 'Fifties', 'Hundreds']
    
    for col in bat_cols:
        if col not in df_bat.columns:
            # Use 0 for numeric fields, '-' for string fields
            if col in ['HS', 'Not_out_status', 'Full_Name']:
                df_bat[col] = '-'
            else:
                df_bat[col] = 0

    numeric_cols = ['Mat', 'Inns', 'NO', 'Runs', 'HS_Numeric', 'BF',
                    'Fours', 'Sixes', 'Ducks', 'Fifties', 'Hundreds']
    for col in numeric_cols:
        df_bat[col] = pd.to_numeric(df_bat[col], errors='coerce').fillna(0)
    
    # Ensure Player_ID is numeric/integer for aggregation later
    df_bat['Player_ID'] = pd.to_numeric(df_bat['Player_ID'], errors='coerce').fillna(0).astype(int)

    # Grouping by Player (short name)
    df_bat_grouped = df_bat.groupby('Player', as_index=False).agg({
        'Player_ID': 'first', # Keep one Player_ID per player
        'Full_Name': 'first', # Keep the Full_Name
        'Season': lambda x: x.nunique(),
        'Mat': 'sum', 'Inns': 'sum', 'NO': 'sum', 'Runs': 'sum',
        'HS': 'first', 
        'HS_Numeric': 'max', # Use max for best HS
        'Not_out_status': 'first',
        'Ave': lambda x: pd.to_numeric(x, errors='coerce').mean(skipna=True), # Use mean of seasonal Ave
        'BF': 'sum', 'SR': lambda x: pd.to_numeric(x, errors='coerce').mean(skipna=True), # Use mean of seasonal SR
        'Hundreds': 'sum', 'Fifties': 'sum', 'Ducks': 'sum',
        'Fours': 'sum', 'Sixes': 'sum'
    })
    df_bat_grouped = df_bat_grouped.rename(columns={'Season': 'Seasons_Played'})
    df_bat_grouped['Player_ID'] = df_bat_grouped['Player_ID'].replace(0, pd.NA).combine_first(df_bat_grouped['Player_ID'])


    # ----------------------------- 🟢 2. BOWLING -----------------------------
    print("Processing Bowling data...")
    df_bowl = read_all_sheets(bowling_path, dtype=str)
    
    # Explicitly define columns and ensure they exist
    bowl_cols = ['Player_ID', 'Player', 'Full_Name', 'Season', 'Mat', 'Inns', 'Overs', 'Mdns', 'Runs', 'Wkts',
                 'BBI', 'Ave', 'Econ', 'SR', '4_Wkts', '5_Wkts']
    
    for col in bowl_cols:
        if col not in df_bowl.columns:
            df_bowl[col] = "0" if col not in ['Full_Name'] else '-'

    # Convert numeric safely
    numeric_cols = ['Mat', 'Inns', 'Overs', 'Mdns', 'Runs', 'Wkts', 'Ave', 'Econ', 'SR', '4_Wkts', '5_Wkts']
    for col in numeric_cols:
        df_bowl[col] = pd.to_numeric(df_bowl[col], errors='coerce').fillna(0)
    
    df_bowl['Player_ID'] = pd.to_numeric(df_bowl['Player_ID'], errors='coerce').fillna(0).astype(int)

    # Ensure BBI is clean and string type
    df_bowl['BBI'] = df_bowl['BBI'].astype(str).str.strip()
    df_bowl['BBI'] = df_bowl['BBI'].apply(lambda x: x if '/' in x else '-')

    def best_bbi(series):
        """Return best BBI (highest wickets, then lowest runs)."""
        valid = [s for s in series if isinstance(s, str) and '/' in s]
        if not valid:
            return '-'
        try:
            parsed = [(int(s.split('/')[0]), int(s.split('/')[1])) for s in valid]
            best = sorted(parsed, key=lambda x: (-x[0], x[1]))[0]
            return f"{best[0]}/{best[1]}"
        except Exception:
            return valid[0]

    df_bowl_grouped = df_bowl.groupby(['Player'], as_index=False).agg({
        'Player_ID': 'first', # Keep one Player_ID per player
        'Full_Name': 'first', # Keep the Full_Name
        'Season': lambda x: x.nunique(),
        'Mat': 'sum', 'Inns': 'sum', 'Overs': 'sum', 'Mdns': 'sum', 'Runs': 'sum',
        'Wkts': 'sum',
        'BBI': best_bbi,
        'Ave': lambda x: pd.to_numeric(x, errors='coerce').mean(skipna=True),
        'Econ': lambda x: pd.to_numeric(x, errors='coerce').mean(skipna=True),
        'SR': lambda x: pd.to_numeric(x, errors='coerce').mean(skipna=True),
        '4_Wkts': 'sum', '5_Wkts': 'sum'
    })

    df_bowl_grouped = df_bowl_grouped.rename(columns={
        'Season': 'Seasons_Played',
        'Mat': 'Mat_bowl', 'Inns': 'Inns_bowl', 'Runs': 'Runs_bowl',
        'Ave': 'Ave_bowl', 'SR': 'SR_bowl'
    })

    # Force text format for BBI
    df_bowl_grouped['BBI'] = df_bowl_grouped['BBI'].astype(str)

    # ----------------------------- 🔵 3. FIELDING -----------------------------
    print("Processing Fielding data...")
    df_field = read_all_sheets(fielding_path)
    
    # Explicitly define columns and ensure they exist
    field_cols = ['Player_ID', 'Player', 'Full_Name', 'Season', 'Mat_field', 'Inns_field', 'Dis', 'Ct', 'St', 'Ct_Wk', 'Ct_Fi', 'MD', 'D/I']
    
    for col in field_cols:
        if col not in df_field.columns:
            df_field[col] = 0 if col not in ['Full_Name'] else '-'

    # Create a temporary column to hold the full MD string value.
    df_field['MD_Full_String'] = df_field['MD'].astype(str)

    # Extract the numeric value for summation (this is correct for grouping)
    df_field['MD_Numeric'] = df_field['MD'].astype(str).str.extract(r'(\d+)').fillna(0)
    df_field['MD_Numeric'] = pd.to_numeric(df_field['MD_Numeric'], errors='coerce').fillna(0)

    num_cols_field = ['Mat_field', 'Inns_field', 'Dis', 'Ct', 'St', 'Ct_Wk', 'Ct_Fi', 'D/I']
    for col in num_cols_field:
        df_field[col] = pd.to_numeric(df_field[col], errors='coerce').fillna(0)
    
    df_field['Player_ID'] = pd.to_numeric(df_field['Player_ID'], errors='coerce').fillna(0).astype(int)

    # Function to get the most detailed MD string (or the first non-empty one)
    def get_max_md_string(series):
        valid = [s for s in series if s not in ['0', 'nan', '-', '0.0'] and '(' in s]
        if valid:
            return max(valid, key=len)
        valid_fallback = [s for s in series if s not in ['0', 'nan', '-', '0.0']]
        if valid_fallback:
            return valid_fallback[0]
        return '-'

    df_field_grouped = df_field.groupby(['Player'], as_index=False).agg({
        'Player_ID': 'first', # Keep one Player_ID per player
        'Full_Name': 'first', # Keep the Full_Name
        'Season': lambda x: x.nunique(),
        'Mat_field': 'sum', 'Inns_field': 'sum', 'Dis': 'sum', 'Ct': 'sum', 'St': 'sum',
        'Ct_Wk': 'sum', 'Ct_Fi': 'sum', 
        'MD_Numeric': 'sum', # Aggregate the numeric value
        'MD_Full_String': get_max_md_string, # Keep the best representative string
        'D/I': 'mean'
    })
    df_field_grouped = df_field_grouped.rename(columns={'Season': 'Seasons_Played', 'MD_Full_String': 'MD_String'})
    # Drop the original 'MD' column which was a mix of string/numeric before grouping
    df_field_grouped = df_field_grouped.drop(columns=['MD_Numeric'], errors='ignore').rename(columns={'MD_String': 'MD'})

    # ----------------------------- 🟣 4. SPAN YEARS -----------------------------
    print("Calculating Player Spans...")
    all_seasons = []
    for path in [batting_path, bowling_path, fielding_path]:
        excel = pd.ExcelFile(path)
        for sheet in excel.sheet_names:
            df = pd.read_excel(path, sheet_name=sheet, dtype=str)
            df.columns = df.columns.str.strip().str.replace(' ', '_', regex=False)
            df["Season"] = sheet.split("_")[-1]
            all_seasons.append(df[["Player", "Season"]])

    if not all_seasons:
        print("⚠️ Warning: No seasonal data found for span calculation.")
        span = pd.DataFrame(columns=["Player", "min", "max", "Span_Years"])
    else:
        df_span = pd.concat(all_seasons, ignore_index=True).drop_duplicates()
        df_span["Season"] = pd.to_numeric(df_span["Season"], errors="coerce")
        span = df_span.groupby("Player")["Season"].agg(["min", "max"]).reset_index()
        span["Span_Years"] = span["min"].fillna(0).astype(int).astype(str) + "-" + span["max"].fillna(0).astype(int).astype(str)
        # Handle single year span case
        span.loc[span['min'] == span['max'], 'Span_Years'] = span['min'].astype(int).astype(str)

    # Merge Span_Years into individual grouped sheets
    df_bat_grouped = pd.merge(df_bat_grouped, span[['Player', 'Span_Years']], on='Player', how='left')
    df_bowl_grouped = pd.merge(df_bowl_grouped, span[['Player', 'Span_Years']], on='Player', how='left')
    df_field_grouped = pd.merge(df_field_grouped, span[['Player', 'Span_Years']], on='Player', how='left')
    
    # ----------------------------- ⚫ 5. MERGE ALL -----------------------------
    print("Merging all data frames...")
    # Base merge using Player, including Full_Name from Batting (will be overwritten if empty)
    df_final = pd.merge(
        df_bat_grouped.drop(columns=['Seasons_Played', 'Player_ID', 'Span_Years'], errors='ignore'), 
        df_bowl_grouped.drop(columns=['Seasons_Played', 'Player_ID', 'Full_Name', 'Span_Years'], errors='ignore'), # Drop duplicate columns
        on="Player", 
        how="outer", 
        suffixes=('_bat', '_bowl')
    )
    
    # Merge with Fielding
    df_final = pd.merge(
        df_final, 
        df_field_grouped.drop(columns=['Seasons_Played', 'Player_ID', 'Full_Name', 'Span_Years'], errors='ignore'), # Drop duplicate columns
        on="Player", 
        how="outer", 
        suffixes=('_bat_bowl', '_field')
    )
    
    # Merge Span, Seasons, and Player_ID (master)
    df_final = pd.merge(df_final, span[['Player', 'Span_Years']], on='Player', how='left')
    
    # Consolidate Seasons_Played (using max from all three dataframes)
    df_seasons = pd.concat([
        df_bat_grouped[['Player', 'Seasons_Played']].rename(columns={'Seasons_Played': 'SP_bat'}),
        df_bowl_grouped[['Player', 'Seasons_Played']].rename(columns={'Seasons_Played': 'SP_bowl'}),
        df_field_grouped[['Player', 'Seasons_Played']].rename(columns={'Seasons_Played': 'SP_field'})
    ]).groupby('Player').max().reset_index()
    df_seasons['Seasons_Played'] = df_seasons[['SP_bat', 'SP_bowl', 'SP_field']].max(axis=1)
    df_final = pd.merge(df_final, df_seasons[['Player', 'Seasons_Played']], on='Player', how='left')
    
    # Consolidate Player_ID (using max, which ignores NA and finds the highest non-zero ID)
    df_ids = pd.concat([
        df_bat_grouped[['Player', 'Player_ID']],
        df_bowl_grouped[['Player', 'Player_ID']],
        df_field_grouped[['Player', 'Player_ID']]
    ]).groupby('Player')['Player_ID'].max().reset_index()
    
    df_final = pd.merge(df_final, df_ids.rename(columns={'Player_ID': 'Player_ID_master'}), on='Player', how='left')
    df_final['Player_ID'] = df_final['Player_ID_master'].fillna(0).astype(int)
    df_final = df_final.drop(columns=['Player_ID_master'], errors='ignore')

    # Final cleanup and NaN replacement
    df_final = df_final.rename(columns={'Full_Name_bat': 'Full_Name'}, errors='ignore') # Ensure Full_Name is consistently named

    for col in df_final.select_dtypes(include='number').columns:
        df_final[col] = df_final[col].fillna(0)
    for col in df_final.select_dtypes(include='object').columns:
        df_final[col] = df_final[col].fillna('-')

    # Column order for Combined Sheet
    final_order = [
        'Player_ID', 'Player', 'Full_Name', 'Span_Years', 'Seasons_Played',
        # Batting
        'Mat', 'Inns', 'NO', 'Runs', 'HS', 'HS_Numeric', 'Not_out_status',
        'Ave', 'BF', 'SR', 'Fours', 'Sixes', 'Ducks', 'Fifties', 'Hundreds',
        # Bowling
        'Mat_bowl', 'Inns_bowl', 'Overs', 'Mdns', 'Runs_bowl', 'Wkts', 'BBI', 'Ave_bowl', 'Econ', 'SR_bowl', '4_Wkts', '5_Wkts',
        # Fielding
        'Mat_field', 'Inns_field', 'Dis', 'Ct', 'St', 'Ct_Wk', 'Ct_Fi', 'MD', 'D/I' # MD is the string column
    ]
    # Reindex columns, dropping any that might not exist after cleaning/merging
    df_final = df_final.reindex(columns=[col for col in final_order if col in df_final.columns])

    # ----------------------------- 🧾 6. SAVE FILES -----------------------------
    # Define preferred column order for individual sheets to place Span_Years and Full_Name early
    def get_sheet_order(df):
        base_cols = ['Player_ID', 'Player', 'Full_Name', 'Span_Years', 'Seasons_Played']
        return base_cols + [c for c in df.columns if c not in base_cols]

    with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
        
        # Prepare and save Batting sheet
        df_bat_grouped.reindex(columns=get_sheet_order(df_bat_grouped)).to_excel(writer, index=False, sheet_name="Batting")

        # Prepare and save Bowling sheet (and rename for standalone view)
        df_bowl_temp = df_bowl_grouped.rename(columns={'Mat_bowl': 'Mat', 'Inns_bowl': 'Inns', 'Runs_bowl': 'Runs', 'Ave_bowl': 'Ave', 'SR_bowl': 'SR'}).copy()
        df_bowl_temp.reindex(columns=get_sheet_order(df_bowl_temp)).to_excel(writer, index=False, sheet_name="Bowling")
        
        # Prepare and save Fielding sheet (Ensure the final MD is the string one)
        df_field_temp = df_field_grouped.rename(columns={'Mat_field': 'Mat', 'Inns_field': 'Inns'}).copy()
        df_field_temp.reindex(columns=get_sheet_order(df_field_temp)).to_excel(writer, index=False, sheet_name="Fielding")
        
        # Save Combined sheet
        df_final.to_excel(writer, index=False, sheet_name="Combined")

    # Format BBI and MD columns as text using openpyxl
    wb = load_workbook(output_excel)
    
    # List of sheet/column pairs to format as text
    text_formats = [
        ("Bowling", "BBI"), ("Combined", "BBI"),
        ("Fielding", "MD"), ("Combined", "MD")
    ]
    
    for sheet_name, col_header in text_formats:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Find the column
            for col_idx in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col_idx).value
                if header == col_header:
                    col_letter = get_column_letter(col_idx)
                    # Apply text format to the entire column
                    for row in range(2, ws.max_row + 1):
                        ws[f"{col_letter}{row}"].number_format = "@"
                    break
    
    wb.save(output_excel)
    print(f"\n✅ All records successfully combined & saved as {output_excel}")

combine_all_records()