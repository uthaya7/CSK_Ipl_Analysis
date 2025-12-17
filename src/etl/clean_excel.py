import openpyxl
from os import path
from openpyxl.utils import get_column_letter
import re
import math # Used for safety checks

# --- CONFIGURATION: FILE PATHS ---
# CHANGED: Input and output paths now use the 'final_excel' directory.
INPUT_DIR = 'F:\\Data Analytics\\Projects\\csk_analysis\\data\\raw_temp'
OUTPUT_DIR = 'F:\\Data Analytics\\Projects\\csk_analysis\\data\\final_excel'

FILE_PATHS = {
    'Batting': path.join(INPUT_DIR, 'batting_records_csk.xlsx'),
    'Bowling': path.join(INPUT_DIR, 'bowling_records_csk.xlsx'),
    'Fielding': path.join(INPUT_DIR, 'fielding_records_csk.xlsx')
}

# --- BATTING CONFIG ---
BATTING_RENAMES = {
    '100': 'Hundreds',
    '50': 'Fifties',
    '0': 'Ducks',
    '4s': 'Fours',
    '6s': 'Sixes'
}
BATTING_COLS_TO_MOVE_TO_END = ['Ducks', 'Fifties', 'Hundreds']
BATTING_COLS_TO_DELETE = ['col15']

# --- BOWLING CONFIG ---
BOWLING_RENAMES = {
    '4': '4_Wkts',
    '5': '5_Wkts'
}
BOWLING_COLS_TO_MOVE_TO_END = ['4_Wkts', '5_Wkts']
BOWLING_COLS_TO_DELETE = ['col14']

# --- FIELDING CONFIG ---
FIELDING_RENAMES = {
    'Ct Wk': 'Ct_Wk',
    'Ct Fi': 'Ct_Fi',
    'Mat': 'Mat_field', # Renaming Mat to avoid merge conflicts later
    'Inns': 'Inns_field' # Renaming Inns to avoid merge conflicts later
}
FIELDING_COLS_TO_DELETE = ['col11']

# --- GLOBAL PLAYER ID & FULL NAME CONFIGURATION ---
GLOBAL_PLAYER_ID_MAP = {}
CURRENT_PLAYER_ID_COUNTER = 1

FULL_PLAYER_NAMES = {
    'A Flintoff': 'Andrew Flintoff', 'A Kamboj': 'Akash Kamboj', 'A Mhatre': 'Aniruddha Mhatre',
    'A Mukund': 'Abhinav Mukund', 'A Nehra': 'Ashish Nehra', 'AM Rahane': 'Ajinkya Madhukar Rahane',
    'AS Rajpoot': 'Ankit Singh Rajpoot', 'AT Rayudu': 'Ambati Tirupati Rayudu', 'Akash Singh': 'Akash Singh',
    'B Laughlin': 'Ben Laughlin', 'BA Stokes': 'Ben Andrew Stokes', 'BB McCullum': 'Brendon Barrie McCullum',
    'BW Hilfenhaus': 'Ben Warwick Hilfenhaus', 'CH Morris': 'Christopher Henry Morris', 'CJ Jordan': 'Christopher James Jordan',
    'CK Kapugedera': 'Chamara Kantha Kapugedera', 'D Brevis': 'Dewald Brevis', 'D Pretorius': 'Dwaine Pretorius',
    'DE Bollinger': 'Douglas Erwin Bollinger', 'DJ Bravo': 'Dwayne John Bravo', 'DJ Hooda': 'Deepak Jagbir Hooda',
    'DJ Hussey': 'David John Hussey', 'DJ Mitchell': 'Daryl Joseph Mitchell', 'DJ Willey': 'David Jonathan Willey',
    'DL Chahar': 'Deepak Lokendrasingh Chahar', 'DP Conway': 'Devon Philip Conway', 'DP Nannes': 'Dirk Peter Nannes',
    'DR Shorey': 'Dhruv Ravinder Shorey', 'DR Smith': 'Dwayne Romell Smith', 'F du Plessis': 'Francois "Faf" du Plessis',
    'GJ Bailey': 'George John Bailey', 'Harbhajan Singh': 'Harbhajan Singh', 'IC Pandey': 'Ishwar Chandra Pandey',
    'Imran Tahir': 'Imran Tahir', 'J Overton': 'Jamie Overton', 'JA Morkel': 'Jacobus "Albie" Morkel',
    'JDP Oram': 'Jacob David Philip Oram', 'JM Kemp': 'Justin Miles Kemp', 'JO Holder': 'Jason Omar Holder',
    'JR Hazlewood': 'Josh Reginald Hazlewood', 'JW Hastings': 'John Wayne Hastings', 'Joginder Sharma': 'Joginder Sharma',
    'KK Ahmed': 'Khizar Khaleel Ahmed', 'KM Asif': 'K. M. Asif', 'KM Jadhav': 'Kedar Mahadev Jadhav',
    'KMDN Kulasekara': 'Kodikara Mudiyanselage Nuwan Kulasekara', 'KV Sharma': 'Karn Vinod Sharma', 'L Balaji': 'Lakshmipathy Balaji',
    'L Ngidi': 'Lungi Ngidi', 'M Manhas': 'Mithun Manhas', 'M Muralidaran': 'Muttiah Muralidaran',
    'M Ntini': 'Makhaya Ntini', 'M Pathirana': 'Matheesha Pathirana', 'M Theekshana': 'Maheesh Theekshana',
    'M Vijay': 'Murali Vijay', 'MA Wood': 'Mark Andrew Wood', 'MEK Hussey': 'Michael Edward Killeen Hussey',
    'MJ Santner': 'Mitchell Josef Santner', 'ML Hayden': 'Matthew Lawrence Hayden', 'MM Ali': 'Moeen Munir Ali',
    'MM Sharma': 'Mohit Mahipal Sharma', 'MS Dhoni': 'Mahendra Singh Dhoni', 'MS Gony': 'Manpreet Singh Gony',
    'Monu Kumar': 'Monu Kumar', 'Mukesh Choudhary': 'Mukesh Choudhary', 'Mustafizur Rahman': 'Mustafizur Rahman',
    'N Jagadeesan': 'Narayan Jagadeesan', 'Noor Ahmad': 'Noor Ahmad Lakanwal', 'P Amarnath': 'P Amarnath',
    'P Negi': 'Pawan Negi', 'PA Patel': 'Parthiv Ajay Patel', 'PH Solanki': 'Prashant Harikisan Solanki',
    'PP Chawla': 'Piyush Pramod Chawla', 'R Ashwin': 'Ravichandran Ashwin', 'R Ravindra': 'Rachin Ravindra',
    'RA Jadeja': 'Ravindrasinh Anirudhsinh Jadeja', 'RA Tripathi': 'Rahul Ajay Tripathi', 'RD Gaikwad': 'Ruturaj Dashrat Gaikwad',
    'RG More': 'Rituraj Gajanan More', 'RJ Gleeson': 'Richard James Gleeson', 'RS Hangargekar': 'Rajvardhan Santosh Hangargekar',
    'RV Uthappa': 'Robin Venu Uthappa', 'S Anirudha': 'Srikkanth Anirudha', 'S Badree': 'Samuel Badree',
    'S Badrinath': 'Subramaniam Badrinath', 'S Dube': 'Shivam Dube', 'S Randiv': 'Suraj Randiv',
    'S Tyagi': 'Shardul Thakur', 'S Vidyut': 'Suresh Vidyut', 'SB Jakati': 'Shadab Bashir Jakati',
    'SB Styris': 'Scott Bernard Styris', 'SC Kuggeleijn': 'Scott Christopher Kuggeleijn', 'SK Raina': 'Suresh Kumar Raina',
    'SK Rasheed': 'Sk Rasheed', 'SM Curran': 'Samuel Matthew Curran', 'SN Thakur': 'Shardul Narendra Thakur',
    'SP Fleming': 'Stephen Paul Fleming', 'SR Watson': 'Shane Robert Watson', 'SSB Magala': 'Sisanda Sandile Bruce Magala',
    'SW Billings': 'Samuel William Billings', 'Sameer Rizvi': 'Sameer Rizvi', 'Simarjeet Singh': 'Simarjeet Singh',
    'T Thushara': 'Thilan Thushara', 'TG Southee': 'Timothy Grant Southee', 'TU Deshpande': 'Tushar Uday Deshpande',
    'UM Patel': 'Urvil Manish Patel', 'V Shankar': 'Vijay Shankar', 'VY Mahesh': 'Veerabhadran Yele Mahesh',
    'WP Saha': 'Wriddhiman Prasanta Saha'
}
# -------------------------------------


def _recalculate_col_map(ws):
    """Generates a mapping of header name to column index (1-based)."""
    header = [str(cell.value).strip() if cell.value is not None else None for cell in ws[1]]
    col_map = {h: i + 1 for i, h in enumerate(header) if h and str(h) != 'None'}
    return col_map

def _delete_empty_columns(ws):
    """Deletes columns that are entirely empty from row 2 onwards."""
    print("    - Checking for and deleting empty columns...")
    cols_deleted = 0
    # Iterate backwards to avoid index shifting issues
    for col_index in range(ws.max_column, 0, -1):
        is_empty = True
        # Check rows 2 up to max_row
        for row_index in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_index, column=col_index).value
            if cell_value is not None and str(cell_value).strip() != '':
                is_empty = False
                break
        
        if is_empty:
            ws.delete_cols(col_index)
            cols_deleted += 1
    
    if cols_deleted > 0:
        print(f"    - Deleted {cols_deleted} empty column(s).")
    return cols_deleted


def _move_columns_to_end(ws, col_names_to_move):
    """Moves a list of columns to the end of the sheet."""
    col_map = _recalculate_col_map(ws)
    
    columns_data = {}
    cols_to_delete_indices = []
    
    for col_name in col_names_to_move:
        if col_name in col_map:
            col_index = col_map[col_name]
            cols_to_delete_indices.append(col_index)
            
            # Extract header and all cell values (including header row 1)
            columns_data[col_name] = [ws.cell(row=r, column=col_index).value for r in range(1, ws.max_row + 1)]
    
    cols_to_delete_indices.sort(reverse=True)
    
    # Delete columns from the old location
    for col_index in cols_to_delete_indices:
        ws.delete_cols(col_index)
    
    # Insert new columns at the end and populate them
    new_max_col = ws.max_column
    for col_name in col_names_to_move:
        if col_name in columns_data:
            new_max_col += 1
            ws.insert_cols(new_max_col) 
            
            data = columns_data[col_name]
            for r, value in enumerate(data):
                ws.cell(row=r + 1, column=new_max_col, value=value)
            
            print(f"    - Moved column '{col_name}' to the end ({get_column_letter(new_max_col)}).")


def _process_player_ids_and_full_name(ws):
    """
    Inserts 'Player_ID' (B), 'Full_Name' (C), and assigns unique IDs globally.
    """
    global GLOBAL_PLAYER_ID_MAP
    global CURRENT_PLAYER_ID_COUNTER
    
    col_map = _recalculate_col_map(ws)
    player_col_index = col_map.get('Player')
    
    if player_col_index:
        # 1. Insert Player_ID column right before Player (to be the new Column B)
        ws.insert_cols(player_col_index)
        ws.cell(row=1, column=player_col_index, value="Player_ID")
        
        # Recalculate map and get the new Player index
        col_map = _recalculate_col_map(ws)
        player_col_index = col_map.get('Player') # Now guaranteed to be one more than Player_ID
        
        # 2. Insert Full_Name column right after Player (to be the new Column C)
        ws.insert_cols(player_col_index + 1)
        ws.cell(row=1, column=player_col_index + 1, value="Full_Name")

        # 3. Assign and fill IDs and Full Names
        for row in range(2, ws.max_row + 1):
            # Player is now at (player_col_index)
            player_name = str(ws.cell(row=row, column=player_col_index).value).strip()
            
            player_id = "N/A"
            full_name = "N/A"

            if player_name and player_name != 'None':
                # --- ID Logic ---
                if player_name not in GLOBAL_PLAYER_ID_MAP:
                    new_id = f"{CURRENT_PLAYER_ID_COUNTER:03d}"
                    GLOBAL_PLAYER_ID_MAP[player_name] = new_id
                    CURRENT_PLAYER_ID_COUNTER += 1
                
                player_id = GLOBAL_PLAYER_ID_MAP[player_name]
                
                # --- Full Name Logic ---
                full_name = FULL_PLAYER_NAMES.get(player_name, player_name) # Fallback to short name
            
            # Write to sheet
            ws.cell(row=row, column=player_col_index - 1, value=player_id) # Player_ID is one column before Player
            ws.cell(row=row, column=player_col_index + 1, value=full_name) # Full_Name is one column after Player

        print("    - Inserted 'Player_ID' and 'Full_Name' columns and assigned unique global IDs.")
        return True
    else:
        print("    - WARNING: 'Player' column not found. Skipping ID/Name column insertion.")
        return False


# --- 1. BATTING TRANSFORMATION FUNCTION (MODIFIED) ---
def transform_batting_sheets(file_path):
    """Loads the workbook, transforms all sheets, and saves the file."""
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"❌ Error: Batting file not found at {file_path}")
        return

    print(f"\n--- BATTING: Transforming '{path.basename(file_path)}' ---")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")

        # --- 1. IDENTIFY, RENAME, AND DELETE COLUMNS ---
        col_map = _recalculate_col_map(ws)
        
        # A. Delete unwanted columns (col15)
        for col_name in BATTING_COLS_TO_DELETE:
            if col_name in col_map:
                col_to_delete_index = col_map[col_name]
                ws.delete_cols(col_to_delete_index)
                col_map = _recalculate_col_map(ws) 
                print(f"    - Removed column '{col_name}'.")

        # B. Rename columns
        for old_name, new_name in BATTING_RENAMES.items():
            if old_name in col_map:
                col_index = col_map[old_name]
                ws.cell(row=1, column=col_index, value=new_name)
                print(f"    - Renamed '{old_name}' to '{new_name}'.")
        
        col_map = _recalculate_col_map(ws)
        
        # --- NEW STEP: INSERT PLAYER_ID AND FULL_NAME (Columns B and D) ---
        # NOTE: Player column is assumed to be A, so Player_ID is B, Player is C, Full_Name is D
        _process_player_ids_and_full_name(ws)
        col_map = _recalculate_col_map(ws) 
        
        # --- EXTRACT AND INSERT SEASON COLUMN (Column E) ---
        season_match = re.search(r'\d{4}', sheet_name)
        season_year = int(season_match.group(0) if season_match else 0)
        
        # Player column is now at column C (index 3). Full_Name is D (index 4). Insert Season at E (index 5).
        player_col_index = col_map.get('Player')
        if player_col_index:
            ws.insert_cols(player_col_index + 2) # Insert 2 columns after Player
            season_col_index = player_col_index + 2
            ws.cell(row=1, column=season_col_index, value="Season")
            
            # Fill Season data
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=season_col_index, value=season_year)
            
            col_map = _recalculate_col_map(ws) 
            print("    - Inserted new column 'Season'.")
        else:
            print("    - WARNING: 'Player' column not found. Skipping 'Season' column insertion.")


        # C. Delete empty columns
        _delete_empty_columns(ws)
        col_map = _recalculate_col_map(ws)

        # D. Move columns to end
        _move_columns_to_end(ws, BATTING_COLS_TO_MOVE_TO_END)
        col_map = _recalculate_col_map(ws)

        # --- 3. INSERT HS_Numeric and Not_out_status COLUMNS (New Columns) ---
        hs_col_index = col_map.get('HS')

        if hs_col_index:
            # Insert two new columns after HS.
            ws.insert_cols(hs_col_index + 1, amount=2) 
            
            col_g_index = hs_col_index + 1
            col_h_index = hs_col_index + 2
            
            hs_col_letter = get_column_letter(hs_col_index)
            
            FORMULA_HS_NUMERIC_ADJUSTED = f'=IF({hs_col_letter}2="-", "-", VALUE(SUBSTITUTE({hs_col_letter}2,"*","")))'
            FORMULA_NOT_OUT_STATUS_ADJUSTED = f'=IF({hs_col_letter}2="-", "-", IF(ISNUMBER(FIND("*", {hs_col_letter}2)), "Not Out", "Out"))'
            
            ws.cell(row=1, column=col_g_index, value="HS_Numeric")
            ws.cell(row=1, column=col_h_index, value="Not_out_status")
            print(f"    - Inserted new columns 'HS_Numeric' and 'Not_out_status'.")

            max_row = ws.max_row
            
            for row in range(2, max_row + 1):
                # Column G (HS_Numeric)
                formula_g = FORMULA_HS_NUMERIC_ADJUSTED.replace('2', str(row))
                ws.cell(row=row, column=col_g_index, value=formula_g)

                # Column H (Not_out_status)
                formula_h = FORMULA_NOT_OUT_STATUS_ADJUSTED.replace('2', str(row))
                ws.cell(row=row, column=col_h_index, value=formula_h)
            
            print(f"    - Applied formulas to {max_row - 1} rows, referencing '{hs_col_letter}'.")
        else:
            print("    - WARNING: 'HS' column not found. Skipping formula insertion.")


    # --- 4. SAVE THE WORKBOOK ---
    new_file_path = path.join(OUTPUT_DIR, path.basename(file_path).replace('.xlsx', '_cleaned.xlsx'))
    wb.save(new_file_path)
    print(f"\n🎉 SUCCESS: Batting transformations saved to: {new_file_path}")
    return new_file_path

# --- 2. BOWLING TRANSFORMATION FUNCTION (MODIFIED) ---
def transform_bowling_sheets(file_path):
    """Loads the workbook, transforms all sheets, and saves the file."""
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"❌ Error: Bowling file not found at {file_path}")
        return

    print(f"\n--- BOWLING: Transforming '{path.basename(file_path)}' ---")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")

        col_map = _recalculate_col_map(ws)
        
        # A. Delete unwanted columns (col14)
        for col_name in BOWLING_COLS_TO_DELETE:
            if col_name in col_map:
                col_to_delete_index = col_map[col_name]
                ws.delete_cols(col_to_delete_index)
                col_map = _recalculate_col_map(ws)
                print(f"    - Removed column '{col_name}'.")

        # B. Rename columns ('4' to '4_Wkts', '5' to '5_Wkts')
        for old_name, new_name in BOWLING_RENAMES.items():
            if old_name in col_map:
                col_index = col_map[old_name]
                ws.cell(row=1, column=col_index, value=new_name)
                print(f"    - Renamed '{old_name}' to '{new_name}'.")
        
        col_map = _recalculate_col_map(ws)
        
        # --- NEW STEP: INSERT PLAYER_ID AND FULL_NAME (Columns B and D) ---
        _process_player_ids_and_full_name(ws)
        col_map = _recalculate_col_map(ws)
        
        # --- EXTRACT AND INSERT SEASON COLUMN (Column E) ---
        season_match = re.search(r'\d{4}', sheet_name)
        season_year = int(season_match.group(0) if season_match else 0)
        
        player_col_index = col_map.get('Player')
        if player_col_index:
            ws.insert_cols(player_col_index + 2) # Insert 2 columns after Player
            season_col_index = player_col_index + 2
            ws.cell(row=1, column=season_col_index, value="Season")
            
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=season_col_index, value=season_year)
            
            col_map = _recalculate_col_map(ws) 
            print("    - Inserted new column 'Season'.")


        # C. Delete empty columns
        _delete_empty_columns(ws)
        col_map = _recalculate_col_map(ws)

        # D. Change BBI datatype to text
        bbi_col_index = col_map.get('BBI')
        if bbi_col_index:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=bbi_col_index).number_format = '@' 
            print("    - Set 'BBI' column data type to Text.")
        
        # E. Move columns to end (4_Wkts and 5_Wkts)
        _move_columns_to_end(ws, BOWLING_COLS_TO_MOVE_TO_END)

    # --- SAVE THE WORKBOOK ---
    new_file_path = path.join(OUTPUT_DIR, path.basename(file_path).replace('.xlsx', '_cleaned.xlsx'))
    wb.save(new_file_path)
    print(f"\n🎉 SUCCESS: Bowling transformations saved to: {new_file_path}")
    return new_file_path


# --- 3. FIELDING TRANSFORMATION FUNCTION (MODIFIED) ---
def transform_fielding_sheets(file_path):
    """Loads the workbook, transforms all sheets, and saves the file."""
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print(f"❌ Error: Fielding file not found at {file_path}")
        return

    print(f"\n--- FIELDING: Transforming '{path.basename(file_path)}' ---")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")

        col_map = _recalculate_col_map(ws)
        
        # A. Delete unwanted columns (col11)
        for col_name in FIELDING_COLS_TO_DELETE:
            if col_name in col_map:
                col_to_delete_index = col_map[col_name]
                ws.delete_cols(col_to_delete_index)
                col_map = _recalculate_col_map(ws)
                print(f"    - Removed column '{col_name}'.")

        # B. Rename columns (Ct Wk to Ct_Wk, Ct Fi to Ct_Fi) and common ones (Mat, Inns)
        for old_name, new_name in FIELDING_RENAMES.items():
            if old_name in col_map:
                col_index = col_map[old_name]
                ws.cell(row=1, column=col_index, value=new_name)
                print(f"    - Renamed '{old_name}' to '{new_name}'.")
        
        col_map = _recalculate_col_map(ws)
        
        # --- NEW STEP: INSERT PLAYER_ID AND FULL_NAME (Columns B and D) ---
        _process_player_ids_and_full_name(ws)
        col_map = _recalculate_col_map(ws)
        
        # --- EXTRACT AND INSERT SEASON COLUMN (Column E) ---
        season_match = re.search(r'\d{4}', sheet_name)
        season_year = int(season_match.group(0) if season_match else 0)
        
        player_col_index = col_map.get('Player')
        if player_col_index:
            ws.insert_cols(player_col_index + 2) # Insert 2 columns after Player
            season_col_index = player_col_index + 2
            ws.cell(row=1, column=season_col_index, value="Season")
            
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=season_col_index, value=season_year)
            
            col_map = _recalculate_col_map(ws) 
            print("    - Inserted new column 'Season'.")


        # C. Delete empty columns
        _delete_empty_columns(ws)

    # --- SAVE THE WORKBOOK ---
    new_file_path = path.join(OUTPUT_DIR, path.basename(file_path).replace('.xlsx', '_cleaned.xlsx'))
    wb.save(new_file_path)
    print(f"\n🎉 SUCCESS: Fielding transformations saved to: {new_file_path}")
    return new_file_path


# --- Main Execution Block ---
def run_all_transformations():
    """Executes the cleaning for all three files."""
    
    # 1. Clean Batting File
    batting_file_path = FILE_PATHS['Batting']
    transform_batting_sheets(batting_file_path)

    # 2. Clean Bowling File
    bowling_file_path = FILE_PATHS['Bowling']
    transform_bowling_sheets(bowling_file_path)

    # 3. Clean Fielding File
    fielding_file_path = FILE_PATHS['Fielding']
    transform_fielding_sheets(fielding_file_path)


if __name__ == "__main__":
    print("--- STARTING DATA TRANSFORMATION AND GLOBAL PLAYER ID ASSIGNMENT ---")
    run_all_transformations()
    print("\n--- GLOBAL PLAYER ID SUMMARY ---")
    print(f"Total unique players found: {len(GLOBAL_PLAYER_ID_MAP)}")
    print("Example IDs (Player: ID):")
    # Print the first 5 or fewer entries
    for i, (player, id) in enumerate(GLOBAL_PLAYER_ID_MAP.items()):
        if i < 5:
            print(f"  {player}: {id}")
        else:
            break